from openpyxl import load_workbook
##wbload will load excel file
wbload=load_workbook('C:/Users/gopal/Desktop/Ravgins/Internship Task.xlsm',use_iterators=True,data_only=True)
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

###wb_create for new excel file
wb_create = Workbook()
dest_filename = 'F:\InternDONE.xlsx'

print '###===================================Starting for sheet Factor1.1========================='
wsc1 = wb_create.create_sheet("Factor1.1",0)

wsr1=wbload.worksheets[1]
row_count = wsr1.max_row
column_count = wsr1.max_column
Total=0

for row1 in range(1, 5):
    for col1 in range(1, 3):
		from_cell=wsr1.cell(column=col1, row=row1)
		to_cell=wsc1.cell(column=col1, row=row1)
		to_cell.value=from_cell.value
		for ro in range(3,5) :	
				for col in range(3,column_count):
					Total=Total+wsr1.cell(row=ro,column=col).value
					col=col+1
				wsc1.cell(row=ro, column=3).value=Total
				Total=0
				ro=ro+1
wb_create.save(filename = dest_filename)

print'##=================Sum Value for Shteet Factor1.2 for Coulmn 5 to 52==========================='
wsc2 = wb_create.create_sheet("Factor1.2",1)
#wsc2.title = "Factor1.2"

wsr2=wbload.worksheets[2]
row_count = wsr2.max_row
column_count = wsr2.max_column
Total=0

for row1 in range(4, row_count):
    for col1 in range(4, 5):
		from_cell=wsr2.cell(column=col1, row=row1)
		to_cell=wsc2.cell(column=col1, row=row1)
		to_cell.value=from_cell.value
		for ro in range(4,row_count) :	
				for col in range(5,52):
					Total=Total+wsr2.cell(row=ro,column=col).value
					col=col+1
				wsc2.cell(row=ro, column=3).value=Total
				Total=0
				ro=ro+1
wb_create.save(filename = dest_filename)
print'##======================For Coulmn 109 to 156==============================='
wsc2 = wb_create.create_sheet("Factor1.2",1)

wsr2=wbload.worksheets[2]
row_count = wsr2.max_row
column_count = wsr2.max_column
Total=0

for row1 in range(4, row_count):
    for col1 in range(7, 9):
		from_cell=wsr2.cell(column=col1, row=row1)
		to_cell=wsc2.cell(column=col1, row=row1)
		to_cell.value=from_cell.value
		for ro in range(5,row_count) :	
				for col in range(109,156):
					Total=Total+wsr2.cell(row=ro,column=col).value
					col=col+1
				wsc2.cell(row=ro, column=3).value=Total
				Total=0
				ro=ro+1
wb_create.save(filename = dest_filename)
print'##======================== For column 159 to end============================='

Total=0

for ro in range(5,row_count) :	
	for col in range(159,column_count):
		Total=Total+sheet2.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1


print'##==============Sum Value of each row for sheet Factor 1.3======================'

sheet3=wb.worksheets[3]
row_count = sheet3.max_row
row_count = row_count -3
column_count = sheet3.max_column-10

Total=0

for ro in range(3,row_count) :	
	for col in range(3,column_count):
		Total=Total+sheet3.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1


print'##==============Sum Value of each row for sheet Factor 1.4======================'

sheet4=wb.worksheets[4]
row_count = sheet4.max_row
column_count = sheet4.max_column

Total=0

for ro in range(4,row_count) :	
	for col in range(57,column_count):
		Total=Total+sheet4.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1



print'##=======================Starting for sheet Factor1.5=============================='

sheet5=wb.worksheets[5]
row_count = sheet5.max_row
column_count = sheet5.max_column

Total=0

for ro in range(8,row_count) :	
	for col in range(3,50):
		Total=Total+sheet5.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1


print'##=======================Starting for sheet Factor1.5 from column 105 to end=============================='


Total=0

for ro in range(8,row_count) :	
	for col in range(105,column_count):
		Total=Total+sheet5.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1



print'##=======================Starting for sheet Factor3 from column 3 to 50=============================='

sheet7=wb.worksheets[7]
row_count = sheet7.max_row
column_count = sheet7.max_column

Total=0

for ro in range(8,row_count) :	
	for col in range(3,50):
		Total=Total+sheet7.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1


print'##============================Factor3, from column 105 till last column=========================================='

Total=0

for ro in range(8,row_count) :	
	for col in range(105,column_count):
		Total=Total+sheet5.cell(row=ro,column=col).value
		col=col+1
	print Total
	Total=0
	ro=ro+1


